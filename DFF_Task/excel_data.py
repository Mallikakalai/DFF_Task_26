import openpyxl
class ex_read():
	"""Read Excel and manupulate it"""
	def __init__(self,filename):
		self.file = filename
		self.wb_obj = openpyxl.load_workbook(filename)
	def readdata(self):
		self.sheet_obj = self.wb_obj.active
		self.max_col = self.sheet_obj.max_column
		self.max_row = self.sheet_obj.max_row
		self.dict_oo={}
		for j in range(2, self.max_row+1):
			self.dict_oo[j-1]={}
			for i in range(1, self.max_col+1):
				self.dict_oo[j-1][self.sheet_obj.cell(row = 1, column = i).value]=self.sheet_obj.cell(row = j, column = i).value
		return self.dict_oo
	def updatecellvalue(self,row,col,value):
		self.sheet_obj.cell(row=row, column=col).value = value
		self.wb_obj.save(self.file)